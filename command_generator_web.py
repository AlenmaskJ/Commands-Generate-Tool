import sys
import os
from flask import Flask, render_template_string, request, send_file, flash, redirect, url_for
import re
import tempfile
from werkzeug.utils import secure_filename
# 替换pandas：用轻量的openpyxl/xlrd读取Excel
import openpyxl
import xlrd


# ========== 单文件exe路径适配（核心！） ==========
def resource_path(relative_path):
    """获取单exe模式下的真实文件路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


# ========== Flask初始化 ==========
app = Flask(__name__)
app.secret_key = 'command_generator_2026'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB文件限制

# ========== 前端页面模板 ==========
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>通用命令生成工具</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        .container {max-width: 800px; margin-top: 50px;}
        .result-area {margin-top: 20px; padding: 15px; border: 1px solid #ddd; border-radius: 5px; min-height: 200px; background-color: #f8f9fa; white-space: pre-wrap; font-family: monospace;}
        .form-label {font-weight: bold;}
    </style>
</head>
<body>
    <div class="container">
        <h2 class="text-center mb-4">通用命令生成工具</h2>
        {% with messages = get_flashed_messages() %}
          {% if messages %}
            {% for message in messages %}
              <div class="alert alert-warning alert-dismissible fade show" role="alert">
                {{ message }}
                <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
              </div>
            {% endfor %}
          {% endif %}
        {% endwith %}
        <form method="POST" enctype="multipart/form-data">
            <div class="mb-3">
                <label for="template" class="form-label">命令模板（变量用{变量名}标识，例：Mod Bscex : MSISDN={x},SMSRouterID={y};）</label>
                <textarea class="form-control" id="template" name="template" rows="3" required placeholder="请输入命令模板，如：Mod Bscex : MSISDN={x},SMSRouterID={y};"></textarea>
            </div>
            <div class="mb-3">
                <label for="excel_file" class="form-label">上传Excel文件（列名对应模板中的变量）</label>
                <input class="form-control" type="file" id="excel_file" name="excel_file" accept=".xlsx,.xls" required>
            </div>
            <button type="submit" class="btn btn-primary">生成命令</button>
        </form>
        {% if commands %}
            <h4 class="mt-4">生成结果</h4>
            <div class="result-area" id="commandResult">{{ commands }}</div>
            <div class="mt-3">
                <a href="{{ url_for('download_file', filename=output_file) }}" class="btn btn-success">下载命令文件（TXT）</a>
            </div>
        {% endif %}
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
'''


# ========== 核心函数：提取模板变量 ==========
def extract_variables(template):
    pattern = r'\{(\w+)\}'  # 匹配{变量名}
    variables = re.findall(pattern, template)
    return list(set(variables))


# ========== 核心函数：生成命令（轻量Excel读取） ==========
def generate_commands_from_template(template, excel_path):
    # 提取变量
    variables = extract_variables(template)
    if not variables:
        raise ValueError("命令模板中未检测到有效变量（需用{变量名}格式，如{x}）")

    commands = []
    # 读取xlsx文件
    if excel_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        # 获取表头
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        # 检查变量是否存在
        missing_vars = [var for var in variables if var not in headers]
        if missing_vars:
            raise ValueError(f"Excel缺少变量列：{', '.join(missing_vars)}")
        # 遍历数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            # 跳过空值行
            if any(row_dict.get(var) is None for var in variables):
                continue
            # 生成命令
            command = template
            for var in variables:
                command = command.replace(f'{{{var}}}', str(row_dict[var]))
            commands.append(command)
    # 读取xls文件
    elif excel_path.endswith('.xls'):
        wb = xlrd.open_workbook(excel_path)
        ws = wb.sheet_by_index(0)
        # 获取表头
        headers = [ws.cell_value(0, col) for col in range(ws.ncols) if ws.cell_value(0, col) is not None]
        missing_vars = [var for var in variables if var not in headers]
        if missing_vars:
            raise ValueError(f"Excel缺少变量列：{', '.join(missing_vars)}")
        # 遍历数据行
        for row_idx in range(1, ws.nrows):
            row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
            row_dict = dict(zip(headers, row))
            if any(row_dict.get(var) is None for var in variables):
                continue
            command = template
            for var in variables:
                command = command.replace(f'{{{var}}}', str(row_dict[var]))
            commands.append(command)
    else:
        raise ValueError("仅支持.xlsx/.xls格式的Excel文件")

    if not commands:
        raise ValueError("未生成有效命令，请检查Excel数据")
    return commands


# ========== 保存命令到临时文件 ==========
def save_commands_to_file(commands):
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, "generated_commands.txt")
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(commands))
    return os.path.basename(file_path), file_path


# ========== 路由：主页面 ==========
@app.route('/', methods=['GET', 'POST'])
def index():
    commands = None
    output_file = None
    if request.method == 'POST':
        try:
            template = request.form['template'].strip()
            excel_file = request.files['excel_file']

            if not excel_file.filename:
                flash("请选择Excel文件")
                return redirect(url_for('index'))

            # 保存上传的Excel到临时文件
            temp_excel_path = os.path.join(tempfile.gettempdir(), secure_filename(excel_file.filename))
            excel_file.save(temp_excel_path)

            # 生成命令
            command_list = generate_commands_from_template(template, temp_excel_path)
            commands = '\n'.join(command_list)

            # 保存命令文件
            output_file, _ = save_commands_to_file(command_list)

            # 删除临时Excel
            os.remove(temp_excel_path)
        except ValueError as e:
            flash(f"错误：{str(e)}")
        except Exception as e:
            flash(f"系统错误：{str(e)}")
    return render_template_string(HTML_TEMPLATE, commands=commands, output_file=output_file)


# ========== 路由：下载文件 ==========
@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(tempfile.gettempdir(), filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name="generated_commands.txt", mimetype='text/plain')
    else:
        flash("下载文件不存在")
        return redirect(url_for('index'))


# ========== 启动应用 ==========
if __name__ == '__main__':
    print("===== 通用命令生成工具 =====")
    print("访问地址：http://127.0.0.1:5000")
    print("按 CTRL+C 退出")
    print("============================")
    app.run(host='0.0.0.0', port=5000, debug=False)  # 关闭调试模式，减小体积